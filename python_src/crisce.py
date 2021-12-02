"""
The Crisce class declares the interface that transform a crisce scenario to json files.

"""



from pre_processing import Pre_Processing
from roads import Roads
from car import Car
from kinematics import Kinematics
import sys
import copy
import itertools
import seaborn as sns
import pandas as pd
import time
import glob
import os
import json
import click
import math
import matplotlib.pyplot as plt
import matplotlib.image as mpimg
from matplotlib.transforms import Bbox
import imutils
import numpy as np
import cv2
import scipy.special
import matplotlib
from shapely.geometry import MultiLineString, Polygon
import scipy.interpolate as scipy_interpolate
import xlsxwriter
from openpyxl import load_workbook

from beamngpy import BeamNGpy, Scenario, Road, Vehicle, setup_logging
from beamngpy.sensors import Electrics, Camera, Damage, Timer

from PythonRobotics.PathPlanning.CubicSpline.cubic_spline_planner import Spline2D
from PythonRobotics.PathPlanning.BezierPath.bezier_path import calc_4points_bezier_path, calc_bezier_path, bernstein_poly, bezier, bezier_derivatives_control_points, curvature, plot_arrow
from PythonRobotics.PathPlanning.BSplinePath.bspline_path import approximate_b_spline_path,  interpolate_b_spline_path


# show_animation = True
crash_impact_model = {
    "front_left": [
        "headlight_L",
        "hood",
        "fender_L",
        "bumper_F",
        "bumperbar_F",
        "suspension_F",
        "body_wagon"
    ],

    "front_right": [
        "hood",
        "bumper_F",
        "bumperbar_F",
        "fender_R",
        "headlight_R",
        "body_wagon",
        "suspension_F"
    ],

    "front_mid": [
        "bumperbar_F",
        "radiator",
        "body_wagon",
        "headlight_R",
        "headlight_L",
        "bumper_F",
        "fender_R",
        "fender_L",
        "hood",
        "suspension_F"
    ],

    "left_mid": [
        "door_RL_wagon",
        "body_wagon",
        "doorglass_FL",
        "mirror_L",
        "door_FL"
    ],

    "rear_left": [
        "suspension_R",
        "exhaust_i6_petrol",
        "taillight_L"
        "body_wagon",
        "bumper_R",
        "tailgate",
        "bumper_R"
    ],

    "rear_mid": [
        "tailgate",
        "tailgateglass",
        "taillight_L",
        "taillight_R",
        "exhaust_i6_petrol",
        "bumper_R",
        "body_wagon",
        "suspension_R"
    ],

    "rear_right": [
        "suspension_R",
        "exhaust_i6_petrol",
        "taillight_R",
        "body_wagon",
        "tailgateglass",
        "tailgate",
        "bumper_R"
    ],

    "right_mid": [
        "door_RR_wagon",
        "body_wagon",
        "doorglass_FR",
        "mirror_R",
        "door_FR"
    ]

}

# sns.set()  # Make seaborn set matplotlib styling


### constants

# RED_CAR_BOUNDARY = np.array([[0, 200, 180],   # red internal_1
#                             [110, 255, 255]])

# # RED_CAR_BOUNDARY = np.array([[0, 190, 215],  # red external_0
# #                             [179, 255, 255]])

# BLUE_CAR_BOUNDARY   = np.array([[85, 50, 60],   
#                                 [160, 255, 255]])

# # blue_car_boundary = np.array([[85, 50, 60],   # blue internal_1
# #                              [160, 255, 255]])


# RED_CAR_BOUNDARY = np.array([[0, 180, 180],   # red internal_2
#                             [110, 255, 255]])


# RED_CAR_BOUNDARY = np.array([[0, 190, 215],  # red external_0
#                             [179, 255, 255]])




# RED_CAR_BOUNDARY = np.array([[0, 190, 210],   # red external_1
#                             [179, 255, 255]])


# # red_car_boundary = np.array([[0, 175, 217],   # red internal_1
# #                             [110, 255, 255]])


# red_car_boundary = np.array([[0, 180, 180],   # red internal_2
#                             [110, 255, 255]])

# # blue_car_boundary = np.array([[85, 30, 40],
# #                              [160, 255, 255]])



class Simulation():
    def __init__(self, vehicles, roads, lane_nodes, kinematics, time_efficiency, output_folder,
                 car_length, car_width, car_length_sim, sketch_type_external, height, width):
        self.crash_analysis_log = dict()
        self.crash_analysis_log["vehicles"] = dict()
        self.crash_analysis_log["roads"] = dict()
        self.bng = None
        self.scenario = None
        self.log = dict()
        self.output_folder = output_folder
        self.vehicles = vehicles
        self.roads = roads
        self.lane_nodes = lane_nodes
        self.kinematics = kinematics
        self.time_efficiency = time_efficiency
        self.car_length = car_length
        self.car_width = car_width
        self.car_length_sim = car_length_sim
        self.sketch_type_external = sketch_type_external
        self.height = height
        self.width = width
        self.process_number = 0

    
    def setupBeamngSimulation(self, file):
        t0 = time.time()
        bng_home = os.getenv('BNG_HOME')
        bng_research = os.getenv('BNG_RESEARCH')
        host = '127.0.0.1'
        port = 64257

        beamng = BeamNGpy(host, port, bng_home, bng_research)
        beamng = beamng.open(launch=True)
        scenario = Scenario('smallgrid', "{}".format(file.split("\\")[-2]))  # 'CRISCE')#
        # beamng.set_tod()
        electrics = Electrics()
        damage  = Damage()
        timer   = Timer()
        
        width_of_coverage   = self.width * self.car_length_sim / self.car_length
        distorted_height    = self.height * self.car_length_sim / self.car_length
        lane_midpoints      = self.roads["small_lane_midpoints"]
        orig = lane_midpoints[0][len(lane_midpoints[0]) // 2]
        # cam_pos = (orig[0] - 5 , self.height - orig[1], width_of_coverage - 20)
        cam_pos = (orig[0] - 5, distorted_height - orig[1], width_of_coverage)
        cam_dir = (0, 1, -60)
        # Real World 3D picture of the crash sketch
        cam = Camera(cam_pos, cam_dir, 90, (2048, 2048), near_far=(1, 4000), colour=True)
        scenario.add_camera(cam, 'cam')


        road_id = ['main_road_1', 'main_road_2', 'main_road_3', 'main_road_4', 'main_road_5', 'main_road_6']
        for i, lane in enumerate(self.lane_nodes):
            road = Road('track_editor_C_center', rid=road_id[i], interpolate=True)
            road.nodes.extend(self.lane_nodes[i])
            scenario.add_road(road)
        
        for vehicle_color in self.vehicles:
            self.crash_analysis_log["vehicles"][vehicle_color] = dict()
            script  = self.vehicles[vehicle_color]["trajectories"]["script_trajectory"]
            angle   = self.vehicles[vehicle_color]["vehicle_info"]["0"]["angle_of_car"]
            orig_pos = self.vehicles[vehicle_color]["snapshots"][0]["center_of_car"]
            x = orig_pos[0] * self.car_length_sim / self.car_length
            y = distorted_height - (orig_pos[1] * self.car_length_sim / self.car_length)
            # print("vehicle", vehicle_color, "postition = ", x, y)

            vehicle = Vehicle("{}_vehicle".format(vehicle_color), 
                              model='etk800', 
                              licence="{}_007".format(vehicle_color), 
                              color=str.capitalize(vehicle_color))
            scenario.add_vehicle(vehicle, pos=(round(x, 1), round(y, 1), 0), rot=(0, 0, -angle - 90), rot_quat=None)
            # scenario.add_vehicle(vehicle, pos=( round(script[0]["x"], 3), round(script[0]["y"], 3), 0), rot=(0,0, -angle -90) , rot_quat=None)
            vehicle.attach_sensor('electrics', electrics)
            vehicle.attach_sensor("damage", damage)
            vehicle.attach_sensor("timer", timer)
            self.crash_analysis_log["vehicles"][vehicle_color]["vehicle"] = vehicle
            self.crash_analysis_log["vehicles"][vehicle_color]["rot"] = (0, 0, -angle - 90)

        scenario.make(beamng)
        
        # bng.set_steps_per_second(20) #### One second is equal to sps(int) steps per simulation...
        # bng = beamng.open()
        beamng.load_scenario(scenario)
        # After loading, the simulator waits for further input to actually start
        beamng.start_scenario()
        # print(scenario._get_prefab())
        
        
        scenario.update()
        frames = scenario.render_cameras()
        plt.figure(figsize=(30, 30))
        plt.imshow(np.asarray(frames['cam']['colour'].convert('RGB')))
        plt.savefig(self.output_folder + '{}_aerial_camera.jpg'.format(self.process_number), dpi=250)
        # plt.show()
        plt.close()
        
        
        t1 = time.time()
        print("loaded beamng scenario in =", t1-t0)
        self.time_efficiency["beamng_loaded"] = t1-t0
        # self.bng = bng
        return beamng, scenario


    def aerialViewCamera(self):
        """ Aerial Camera setup for taking picture and monitoring the crash """
        width_of_coverage   = max(self.height, self.width) * self.car_length_sim / self.car_length
        distorted_height    = self.height * self.car_length_sim / self.car_length
        a_ratio             = self.car_length_sim / self.car_length
        orig = [(self.height * a_ratio) // 2, (self.width * a_ratio) // 2]
        # cam_pos = (orig[0] - 10, self.height - orig[1], width_of_coverage - 20)
        cam_pos = (orig[0], distorted_height - orig[1], width_of_coverage)
        cam_dir = (0, 1, -60)
        self.bng.set_free_camera(cam_pos, cam_dir)
        
        # # Real World 3D picture of the crash sketch
        # cam = Camera(cam_pos, cam_dir, 80, (2048, 2048), near_far=(1, 4000), colour=True)
        # self.scenario.add_camera(cam, 'cam')
        # self.scenario.update()
        # frames = self.scenario.render_cameras()
        # plt.figure(figsize=(30, 30))
        # plt.imshow(np.asarray(frames['cam']['colour'].convert('RGB')))
        # plt.savefig(self.output_folder + '{}_aerial_camera.jpg'.format(self.process_number), dpi=150)
        # # plt.show()
        # plt.close()


    def plot_road(self, ax, print_road_values=True):
        # fig = plt.figure(figsize=(20, 17))
        # ax = fig.add_subplot(111)
        
        road_names_list = list(self.bng.get_roads().keys())
        road_geometry_edges = [self.bng.get_road_edges(road) for road in road_names_list]
        for i, road in enumerate(road_geometry_edges):
            self.crash_analysis_log["roads"][road_names_list[i]] = dict()

            left_edge_x  = np.array([e['left'][0] for e in road])
            left_edge_y  = np.array([e['left'][1] for e in road])
            right_edge_x = np.array([e['right'][0] for e in road])
            right_edge_y = np.array([e['right'][1] for e in road])
            simulation_road_width = np.sqrt(np.square(right_edge_x[0] - left_edge_x[0]) + np.square(right_edge_y[0] - left_edge_y[0]))
            x_min = min(left_edge_x.min(), right_edge_x.min())
            x_max = max(left_edge_x.max(), right_edge_x.max())
            y_min = min(left_edge_y.min(), right_edge_y.min())
            y_max = max(left_edge_y.max(), right_edge_y.max())
            simulation_road_length = np.sqrt(np.square(x_max - x_min) + np.square(y_max - y_min))
            if print_road_values:
                print("simulation_road_width = ",  simulation_road_width)
                print("simulation_road_length = ", simulation_road_length)
            # simulation_road.append([simulation_road_width, simulation_road_length])
            self.crash_analysis_log["roads"][road_names_list[i]]["simulation_road_width"]  = simulation_road_width
            self.crash_analysis_log["roads"][road_names_list[i]]["simulation_road_length"] = simulation_road_length

            x_min = x_min - 20   # We add/subtract 10 or 17  from the min/max coordinates to pad
            x_max = x_max + 20   # the area of the plot a bit
            y_min = y_min - 20
            y_max = y_max + 20
            ax.set_aspect('equal', 'datalim')
            # pyplot & bng coordinate systems have different origins
            ax.set_xlim(left=x_max, right=x_min)
            ax.set_ylim(bottom=y_max, top=y_min)  # so we flip them here
            ax.plot(left_edge_x, left_edge_y, 'k-')
            ax.plot(right_edge_x, right_edge_y, 'k-')
            # break
            
        # plt.gca().set_aspect("auto")
        # plt.axis(False)
        # plt.gca().invert_yaxis()
        # plt.gca().invert_xaxis()
        # plt.show()
        # plt.close()
        # plt.savefig("Internal_Validity\internal_road_7.jpg", bbox_inches='tight')
        # fig.savefig("Internal_Validity\internal_road_7.jpg", bbox_inches='tight')


    def getBboxRect(self, vehicle):
        bbox = vehicle.get_bbox()
        boundary_x = [
            bbox['front_bottom_left'][0],
            bbox['front_bottom_right'][0],
            bbox['rear_bottom_right'][0],
            bbox['rear_bottom_left'][0],
        ]
        boundary_y = [
            bbox['front_bottom_left'][1],
            bbox['front_bottom_right'][1],
            bbox['rear_bottom_right'][1],
            bbox['rear_bottom_left'][1],
        ]

        return bbox, boundary_x, boundary_y


    def extractingAngleUsingBbox(self, vehicle):
        bbox = self.getBboxRect(vehicle=vehicle)[0]
        fbl = bbox['front_bottom_left']
        fbr = bbox['front_bottom_right']
        rbr = bbox['rear_bottom_right']
        rbl = bbox['rear_bottom_left']
        ### Extracting the front the node and rear node of the car and using them for extracting the angle using arctan2
        front_node = ((fbl[0] + fbr[0]) / 2, (fbl[1] + fbr[1]) / 2)
        rear_node = ((rbl[0] + rbr[0]) / 2, (rbl[1] + rbr[1]) / 2)

        x = [(rear_node[0]), (front_node[0])]
        y = [(rear_node[1]), (front_node[1])]
        dy = y[1] - y[0]
        dx = x[1] - x[0]

        angle_of_vehicle = math.degrees(math.atan2(dy, dx))
        angle_of_vehicle = (-angle_of_vehicle if angle_of_vehicle < 0 else (-angle_of_vehicle + 360))
        angle_of_vehicle = -angle_of_vehicle + 360
        # print("angle = ", angle_of_vehicle)

        return angle_of_vehicle


    def computeTriangle(self, bbox):
        # format e.g [80.80281829833984, 15.481525421142578,   -0.031248807907104492]
        fbl = bbox['front_bottom_left']
        fbr = bbox['front_bottom_right']
        rbr = bbox['rear_bottom_right']
        rbl = bbox['rear_bottom_left']

        ### Midpoint of the Car
        mid_fbl_rbr = ((fbl[0] + rbr[0]) / 2, (fbl[1] + rbr[1]) / 2)
        mid_fbr_rbl = ((fbr[0] + rbl[0]) / 2, (fbr[1] + rbl[1]) / 2)

        tri_left = ((fbl[0] + mid_fbl_rbr[0]) / 2, (fbl[1] + mid_fbl_rbr[1]) / 2)
        tri_right = ((fbr[0] + mid_fbl_rbr[0]) / 2, (fbr[1] + mid_fbl_rbr[1]) / 2)
        tri_top = ((fbl[0] + fbr[0]) / 2, (fbl[1] + fbr[1]) / 2)

        return (tri_left, tri_right, tri_top)


    def centerOfRect(self, boundary_x, boundary_y):
        midpoint = (
            (boundary_x[0] + boundary_x[1] + boundary_x[2] + boundary_x[3]) / len(boundary_x),
            (boundary_y[0] + boundary_y[1] + boundary_y[2] + boundary_y[3]) / len(boundary_y)
        )
        return midpoint


    def plot_bbox_rect(self, ax, vehicle):
        bbox, boundary_x, boundary_y = self.getBboxRect(vehicle=vehicle)
        midpoint = self.centerOfRect(boundary_x, boundary_y)
        print("midpoint = ", midpoint)

        if vehicle.vid == "red_vehicle":
            ax.fill(boundary_x, boundary_y, 'r-')
        else:
            ax.fill(boundary_x, boundary_y, 'b-')

        triangle_coord = self.computeTriangle(bbox)
        # print(triangle_coord)
        # ax.Polygon(np.array(triangle_coord), closed=False, color="blue", alpha=0.3, fill=True, edgecolor=None)
        poly = plt.Polygon(np.array(triangle_coord), closed=False, color="white", alpha=1, fill=True, edgecolor=None)
        circle = plt.Circle((midpoint[0], midpoint[1]), 0.4, color='green')
        ax.add_patch(circle)
        ax.add_patch(poly)


    def plotSimulationCrashSketch(self):
        fig = plt.figure(figsize=(30, 20))
        plt.gca().set_aspect("auto")
        plt.axis('off')

        self.plot_road(plt.gca())
        # plot_script(plt.gca())

        for vehicle_color in self.vehicles:
            vehicle = self.crash_analysis_log["vehicles"][vehicle_color]["vehicle"]
            self.plot_bbox_rect(plt.gca(), vehicle)
            vehicle.update_vehicle()
            angle = self.extractingAngleUsingBbox(vehicle)
            self.crash_analysis_log["vehicles"][vehicle_color]["initial_position"]  = vehicle.state['pos']
            self.crash_analysis_log["vehicles"][vehicle_color]["initial_direction"] = angle
            self.crash_analysis_log["vehicles"][vehicle_color]["init_bbox"]         = self.getBboxRect(vehicle)[0]
            # self.crash_analysis_log["vehicles"][vehicle_color]["center_of_bbox"]    = self.centerOfRect(self.getBboxRect(vehicle=vehicle)[1], 
            #                                                                                             self.getBboxRect(vehicle=vehicle)[2])
            print(vehicle_color, "vehicle \n\tinitial_positions = ", vehicle.state['pos'])
            print("\tAngle =", angle)
            # self.plot_bbox_rect(plt.gca(), vehicle)

        plt.gca().invert_yaxis()
        plt.gca().invert_xaxis()
        plt.savefig(self.output_folder + '{}_sim_initial_pos_dir.jpg'.format(self.process_number), bbox_inches='tight')
        self.process_number += 1
        # plt.show()
        plt.close()
        
    def eulerToDegree(self, euler):
        return ((euler) / (2 * np.pi)) * 360
    
    def eulerToQuat(self, angle):
        """ Converts an euler angle in degree to quaternion. """
        angle = np.radians(angle)

        cy = np.cos(angle[2] * 0.5)
        sy = np.sin(angle[2] * 0.5)
        cp = np.cos(angle[1] * 0.5)
        sp = np.sin(angle[1] * 0.5)
        cr = np.cos(angle[0] * 0.5)
        sr = np.sin(angle[0] * 0.5)

        w = cr * cp * cy + sr * sp * sy
        x = sr * cp * cy - cr * sp * sy
        y = cr * sp * cy + sr * cp * sy
        z = cr * cp * sy - sr * sp * cy

        return (x, y, z, w)

    def initiateDrive(self):
        for vehicle_color in self.vehicles:
            self.bng.add_debug_line(points = self.vehicles[vehicle_color]["trajectories"]["debug_trajectory"],
                                    point_colors = self.vehicles[vehicle_color]["trajectories"]["point_colors"],
                                    spheres = self.vehicles[vehicle_color]["trajectories"]["spheres"],
                                    sphere_colors = self.vehicles[vehicle_color]["trajectories"]["sphere_colors"],
                                    cling = True,
                                    offset = 0.1)

        # bng.step(10)

        """ AI script must have at least 3 nodes """

        vehicle_red     = self.crash_analysis_log["vehicles"]["red"]["vehicle"]
        vehicle_blue    = self.crash_analysis_log["vehicles"]["blue"]["vehicle"]

        position_red    = list()
        positions_blue  = list()
        red_bbox        = list()
        blue_bbox       = list()
        speed_r         = list()
        speed_b         = list()
        direction_r     = list()
        direction_b     = list()
        sim_time_red    = list()
        sim_time_blue   = list()
        position_red.append(self.crash_analysis_log["vehicles"]["red"]["initial_position"])
        positions_blue.append(self.crash_analysis_log["vehicles"]["blue"]["initial_position"])

        t0 = time.time()
        if(len(self.vehicles["red"]["snapshots"]) > 1):
            vehicle_red.ai_set_mode('manual')
            vehicle_red.ai_set_script(self.vehicles["red"]["trajectories"]["script_trajectory"], cling=False)

        if(len(self.vehicles["blue"]["snapshots"]) > 1):
            vehicle_blue.ai_set_mode('manual')
            vehicle_blue.ai_set_script(self.vehicles["blue"]["trajectories"]["script_trajectory"], cling=False)

        self.bng.set_steps_per_second(120) # 80, 100, 120, 150 
        self.bng.pause()
        self.bng.display_gui_message("Drive Initiated !!!")

        while True:
            # time.sleep(0.1)
            # Synchs the vehicle's "state" variable with the simulator
            vehicle_red.update_vehicle()
            vehicle_blue.update_vehicle()
            red_bbox.append(self.getBboxRect(vehicle_red)[0])
            blue_bbox.append(self.getBboxRect(vehicle_blue)[0])
            red_sensors = self.bng.poll_sensors(vehicle=vehicle_red)
            blue_sensors = self.bng.poll_sensors(vehicle=vehicle_blue)
            wheelspeed_r = red_sensors["electrics"]["wheelspeed"] * 3.6
            wheelspeed_b = blue_sensors["electrics"]["wheelspeed"] * 3.6
            speed_r.append(red_sensors["electrics"]["wheelspeed"])
            speed_b.append(blue_sensors["electrics"]["wheelspeed"])
            direction_r.append(vehicle_red.state['dir'])
            direction_b.append(vehicle_blue.state['dir'])
            position_red.append(vehicle_red.state['pos'])
            positions_blue.append(vehicle_blue.state['pos'])
            sim_time_red.append(red_sensors["timer"]["time"])
            sim_time_blue.append(blue_sensors["timer"]["time"])
            if(red_sensors["damage"]["damage"] != 0 or blue_sensors["damage"]["damage"] != 0):
                print("crashed !!!")
                if (wheelspeed_r < 5 and wheelspeed_b < 5):
                    print("vehicles came to halt")
                    break
            self.bng.step(3)

        self.bng.resume()

        t1 = time.time()
        print("\nBeamNG scenario executed in =", t1-t0)
        self.time_efficiency["beamng_executed"] = t1-t0

        self.crash_analysis_log["vehicles"]["red"]["final_bboxes"]  = red_bbox
        self.crash_analysis_log["vehicles"]["blue"]["final_bboxes"] = blue_bbox
        self.crash_analysis_log["vehicles"]["red"]["simulation_trajectory"]     = position_red
        self.crash_analysis_log["vehicles"]["blue"]["simulation_trajectory"]    = positions_blue
        self.crash_analysis_log["vehicles"]["red"]["rotations"]     = direction_r
        self.crash_analysis_log["vehicles"]["blue"]["rotations"]    = direction_b
        self.crash_analysis_log["vehicles"]["red"]["wheel_speed"]   = speed_r
        self.crash_analysis_log["vehicles"]["blue"]["wheel_speed"]  = speed_b
        self.crash_analysis_log["vehicles"]["red"]["sim_time"]      = sim_time_red
        self.crash_analysis_log["vehicles"]["blue"]["sim_time"]     = sim_time_blue

    def postCrashDamage(self):
        t0 = time.time()
        is_crashed = False
        for vehicle_color in self.vehicles:
            vehicle = self.crash_analysis_log["vehicles"][vehicle_color]["vehicle"]
            vehicle_sensor = self.bng.poll_sensors(vehicle=vehicle)
            if (vehicle_sensor["damage"]["damage"] > 0):
                is_crashed = True
                # crash_analysis_log["vehicles"][vehicle_color]["simulation_damage"] = dict()
                self.crash_analysis_log["vehicles"][vehicle_color]["crashed_happened"] = is_crashed
                self.crash_analysis_log["vehicles"][vehicle_color]["simulation_damage"] = vehicle_sensor["damage"]["part_damage"]
                print(vehicle.vid, "crash happened !!!")
        t1 = time.time()
        # print("total time", t1-t0)
        self.time_efficiency["eval_crash_damage"] = t1-t0
    
    
    """1. Quality and Accuracy of Simulated Crash Impact Sides of the Vehicles """

    def effectAndAccurayOfSimulation(self):
        t0 = time.time()
        self.log["vehicles"] = dict()
        
        for vehicle_color in self.vehicles:
            self.log["vehicles"][vehicle_color] = dict()
            
            int_impact_side = self.vehicles[vehicle_color]["impact_point_details"]["internal_impact_side"]
            ref_impact_side_values  = list(self.vehicles[vehicle_color]["impact_point_details"]["reference_deformations"])
            try:
                sim_impact_side_values  = list(self.crash_analysis_log["vehicles"][vehicle_color]["simulation_damage"].keys())
            except:
                sim_impact_side_values = ["No_damage", "No_damage"]
            
            lenght_of_ref_values    = len(ref_impact_side_values)
            lenght_of_sim_values    = len(sim_impact_side_values)

            sim_impact_side_values  = [sim_side_value.split("_", maxsplit=1)[1] for sim_side_value in sim_impact_side_values]

            crash_damage = list()

            for crash_side in crash_impact_model:
                side_values     = set(crash_impact_model[crash_side])
                sim_side_values = set(sim_impact_side_values)
                difference      = side_values.difference(sim_side_values)
                diff_per        = (1 - len(difference) / len(side_values)) * 100
                crash_damage.append([diff_per, crash_side, side_values, sim_side_values])
                # print("crash side calculation", [diff_per, crash_side, side_values, sim_side_values])

            crash_damage        = np.array(crash_damage)
            print("\nCrash values reported by Beamng: \n", crash_damage[crash_damage[:, 0].argsort()][-2:])
            # print("\n crash values", crash_damage)
            crash_damage        = crash_damage[crash_damage[:, 0].argsort()][-2:]

            for crash_values in crash_damage:

                crash_impact_hits   = crash_values[0]
                sim_impact_side     = crash_values[1]
                ref_side_values     = list(crash_values[2])
                sim_side_values     = list(crash_values[3])
                missed_values       = list(crash_values[2].difference(crash_values[3]))
                out_of_bound        = list(crash_values[3].difference(crash_values[2]))

                if crash_values[1] in ['front_left', 'front_right', 'front_mid']:
                    sim_impact_side = "front"
                elif crash_values[1] in ['rear_left', 'rear_mid', 'rear_right']:
                    sim_impact_side = "rear"
                elif crash_values[1] == "right_mid":
                    sim_impact_side = "right"
                elif crash_values[1] == "left_mid":
                    sim_impact_side = "left"

                if not self.sketch_type_external:
                    crash_impact_side = str.split(int_impact_side, "_")[0]
                    self.vehicles[vehicle_color]["impact_point_details"]["external_impact_side"] = "Internal"
                else:
                    ext_impact_side = self.vehicles[vehicle_color]["impact_point_details"]["external_impact_side"]
                    crash_impact_side = str.lower(ext_impact_side)

                crash_impact_match = crash_values[0]
                crash_impact_hits = len(list(crash_values[2].intersection(crash_values[3]))) # len(sim_side_values) - len(missed_values)
                # crash_impact_hits   = (crash_impact_hits if crash_impact_hits >= 0 else 0)
                crash_impact_miss = len(missed_values)

                if sim_impact_side == crash_impact_side:
                    break

            print("\n")
            print("Vehicle", str.capitalize(vehicle_color))
            print("reference impact points from the crash sketch: ")
            print("\t", ref_side_values)
            print("count of reference impact points: ", len(ref_side_values))

            print("simulation impact points from the sensor damage: ")
            print("\t", sim_impact_side_values)
            print("count of simulation impact points: ", lenght_of_sim_values)


            self.crash_analysis_log["vehicles"][vehicle_color]["crash_side_generic"] = str.capitalize(crash_impact_side)
            self.crash_analysis_log["vehicles"][vehicle_color]["crash_side_specific"]= int_impact_side
            self.crash_analysis_log["vehicles"][vehicle_color]["crash_impact_match"] = crash_impact_match
            self.crash_analysis_log["vehicles"][vehicle_color]["crash_side_match"]   = (sim_impact_side == crash_impact_side)
            self.crash_analysis_log["vehicles"][vehicle_color]["crash_impact_hits"]  = crash_impact_hits
            self.crash_analysis_log["vehicles"][vehicle_color]["crash_impact_miss"]  = crash_impact_miss
            self.crash_analysis_log["vehicles"][vehicle_color]["missed_values"]      = missed_values
            self.crash_analysis_log["vehicles"][vehicle_color]["out_of_bound"]       = out_of_bound
            
            #### ---- For Storing  the log in the excel file for data analysis------- ####
            self.log["vehicles"][vehicle_color]["crash_side_generic"]     = str.capitalize(crash_impact_side)
            self.log["vehicles"][vehicle_color]["crash_side_specific"]    = int_impact_side
            self.log["vehicles"][vehicle_color]["crash_impact_match"]     = crash_impact_match
            self.log["vehicles"][vehicle_color]["crash_side_match"]       = (sim_impact_side == crash_impact_side)
            self.log["vehicles"][vehicle_color]["crash_impact_hits"]      = crash_impact_hits
            self.log["vehicles"][vehicle_color]["crash_impact_miss"]      = crash_impact_miss
            self.log["vehicles"][vehicle_color]["missed_values"]          = missed_values
            self.log["vehicles"][vehicle_color]["out_of_bound"]           = out_of_bound
            self.log["vehicles"][vehicle_color]["ref_side_values"]           = ref_side_values

            print("\nEffectiveness of simulation ")
            print("Hits = ", int(crash_impact_hits))
            print("Miss = ", len(missed_values))
            print("Crash Impact Side by CRISCE      = ", str.capitalize(crash_impact_side))
            if self.sketch_type_external:
                print("External Impact Side             = ", str.capitalize(self.vehicles[vehicle_color]["impact_point_details"]["external_impact_side"]))
            print("Crash Impact Side Observed       = ", str.capitalize(crash_values[1])) # .split("_")[0]))
            # print("Crash Impact Hits Percentage     = ", int((len(sim_side_values) - len(missed_values)) / len(ref_side_values) * 100), "%")
            print("Crash Impact Miss Percentage     = ", int(len(missed_values) / len(ref_side_values) * 100), "%")
            print("Crash Impact Match Succussfull   = ", crash_impact_match, "%")
            print("Crash Impact Side Match          = ", (sim_impact_side == crash_impact_side))
            print("Crash Impact Miss Values         = ", missed_values)
            print("Crash Impact Out-of-bound Values = ", out_of_bound)

        print("\n\n\nACCURACY OF THE SIMULATED IMPACTS")
        impact_acc_red  = 50 * self.crash_analysis_log["vehicles"]["red"]["crash_side_match"]
        impact_acc_blue = 50 * self.crash_analysis_log["vehicles"]["blue"]["crash_side_match"]

        if (impact_acc_red + impact_acc_blue) == 100:
            print("TOTAL MATCH")
            simulated_impact = "TM"
        elif(impact_acc_red + impact_acc_blue) == 50:
            print("PARTIAL MATCH")
            simulated_impact = "PM"
        else:
            print("NO MATCH")
            simulated_impact = "NM"
        
        self.crash_analysis_log["vehicles"][vehicle_color]["simulated_impact"] = simulated_impact

        #### ---- For Storing  the log in the excel file for data analysis------- ####
        self.log["simulated_impact"] = simulated_impact

        t1 = time.time()
        print("\nBeamNG evaluation -- side impact calculated in =", t1-t0)
        self.time_efficiency["eval_side_impact"] = t1-t0


    """ 
    2. Quality of Environment:
        a. Road Geometry: Sketch roads comparison with Simulation roads (widths and lengths of the lanes).
        b. Initial Placement of the vehicles: Comparison of the initial vehicle placement in the sketch to the placement in the simulation.
        c. Orientation of Vehicles: Comparison of the orientation of the vehicles from sketch to simulation.
    """

    def distanceMetricComputation(self, point_1, point_2):
        return np.linalg.norm(point_1 - point_2)
    
    def geometricDifference(self, value_1, value_2):
        return (100 - (value_1 / value_2) * 100)

    ### a. Road Geometry
    
    def computeRoadGeometricSimilarity(self):
        t0 = time.time()
        road_geometry = 0
        # number_of_roads = 0
        for i, road_name in enumerate(self.crash_analysis_log["roads"]):
            sim_road_width  = self.crash_analysis_log["roads"][road_name]["simulation_road_width"]
            sim_road_length = self.crash_analysis_log["roads"][road_name]["simulation_road_length"]
            sketch_road_width   = self.roads["scaled_lane_width"][i]
            sketch_road_length  = self.roads["scaled_lane_length"][i]
            width_similarity, length_similarity = False, False
            
            ### 1st Approach as the simulation values for road uses the Cat-Rom-Mull splines so the values will always 
            ### be greater than the sketch values as the values elongated by the splines to ensures the smooth curves.
            if sim_road_width > sketch_road_width:
                width_similarity = True
            if sim_road_length > sketch_road_length:
                length_similarity = True
            if width_similarity == False or length_similarity == False:
                width_similarity    = math.isclose(sim_road_width, sketch_road_width,   abs_tol = 6)
                length_similarity   = math.isclose(sim_road_length, sketch_road_length, abs_tol = 6)
            
            ### 2st Approach takes the difference between the sketch and simulation and the using geometric difference 
            ### gives us the value for the either lenght or the width of the road.
            # width_difference    = self.distanceMetricComputation(np.array(sim_road_width), np.array(sketch_road_width))
            # width_similarity    = self.geometricDifference(width_difference, sketch_road_width)
            # length_difference   = self.distanceMetricComputation(np.array(sim_road_length), np.array(sketch_road_length))
            # length_similarity   = self.geometricDifference(length_difference, sketch_road_length)
            # road_similarity     = (0.5 * width_similarity + 0.5 * length_similarity) / 100
            
            self.crash_analysis_log["roads"][road_name]["width_similarity"]     = width_similarity
            self.crash_analysis_log["roads"][road_name]["length_similarity"]    = length_similarity
            if  (width_similarity and length_similarity):
                road_similarity = 1
            elif (width_similarity or length_similarity):
                road_similarity = 0.5
            else:
                road_similarity = 0
            self.crash_analysis_log["roads"][road_name]["road_similarity"]      = road_similarity
            road_geometry += road_similarity
            # number_of_roads += 1

            print(road_name)
            print("\tsketch road width      = ", sketch_road_width)
            print("\tsimulation road width  = ",    sim_road_width)
            print("\tsketch road length     = ", sketch_road_length)
            print("\tsimulation road length = ",   sim_road_length)
            print("\troad width similar     =  {} %",  width_similarity)
            print("\troad length similar    =  {} %", length_similarity)
            print("\troad geomerty similar  =  {} %".format(road_similarity * 100))

        ## Calculating the road geometry similarity for the environment
        
        t1 = time.time()
        print("\nBeamNG evaluation -- Road Geometry calculated in =", t1-t0)
        self.time_efficiency["eval_quality_road_geom"] = t1 - t0

        road_similarity = (road_geometry / len(self.crash_analysis_log["roads"]) * 100)
        
        #### ---- For Storing  the log in the excel file for data analysis------- ####
        self.log["road_similarity"] = road_similarity
        
        return road_similarity


    ### b and c. Initial Placement and Orientation of the Vehicles

    def computeVehiclesSimilarity(self):
        """ Compute Red and Blue Vehicle Orientataion and Initial placement Similarity """
        t0 = time.time()
        pos_sim = 0
        distorted_height = self.height * self.car_length_sim / self.car_length
        orient_sim = 0
        for vehicle_color in self.vehicles:
            # script_pos   = self.vehicles[vehicle_color]["trajectories"]["script_trajectory"]
            skt_pos = self.vehicles[vehicle_color]["snapshots"][0]["center_of_car"]
            x = skt_pos[0] * self.car_length_sim / self.car_length
            y = distorted_height - (skt_pos[1] * self.car_length_sim / self.car_length)
            # print("vehicle", vehicle_color, "postition = ", x, y)
            init_skt_pos = [x, y, 0]
            init_skt_dir = [self.vehicles[vehicle_color]["vehicle_info"]["0"]["angle_of_car"]]
            init_sim_pos = self.crash_analysis_log["vehicles"][vehicle_color]["initial_position"]
            init_sim_dir = [self.crash_analysis_log["vehicles"][vehicle_color]["initial_direction"]]

            position_distance       = self.distanceMetricComputation(np.array(init_skt_pos), np.array(init_sim_pos))
            orientation_distance    = self.distanceMetricComputation(np.array(init_skt_dir), np.array(init_sim_dir))
            position_similarity     = (True if (position_distance    < 2) else False)
            orientation_similarity  = (True if (orientation_distance < 2) else False)
            self.crash_analysis_log["vehicles"][vehicle_color]["pos_sim"]    = position_similarity
            self.crash_analysis_log["vehicles"][vehicle_color]["orient_sim"] = orientation_similarity
            pos_sim     += position_similarity
            orient_sim  += orientation_similarity

            print("\n", str.capitalize(vehicle_color), "Vehicle")
            print("\tinit_skt_pos =", init_skt_pos)
            print("\tinit_sim_pos =", init_sim_pos)
            print("\tinit_skt_dir =", init_skt_dir)
            print("\tinit_sim_dir =", init_sim_dir)
            print("\tsketch and simulation vehicle positions are similar    = ", (True if (position_distance < 1) else False))
            print("\tsketch and simulation vehicle orientations are similar = ", (True if (orientation_distance < 1) else False))

        t1 = time.time()
        print("\nBeamNG evaluation -- quality of Vehicle Orientataion and Initial placement calculated in =", t1 - t0)
        self.time_efficiency["eval_quality_orien_pos"] = t1 - t0
        
        #### ---- For Storing  the log in the excel file for data analysis------- ####
        placement_similarity    = pos_sim / len(self.vehicles) * 100
        orientation_similarity  = orient_sim / len(self.vehicles) * 100
        self.log["placement_similarity"]     = placement_similarity
        self.log["orientation_similarity"]   = orientation_similarity

        return (placement_similarity, orientation_similarity)

    def computeSimVehBboxCoord(self, sim_box):
        # sim_box = self.crash_analysis_log["vehicles"][v_color]["init_bbox"]
        distorted_height = self.height * self.car_length_sim / self.car_length
        fbl = sim_box['front_bottom_left'][:2]
        fbr = sim_box['front_bottom_right'][:2]
        rbr = sim_box['rear_bottom_right'][:2]
        rbl = sim_box['rear_bottom_left'][:2]
        sim_veh  = np.asarray([fbl, fbr,  rbr, rbl])
        sim_veh  = np.array([tuple([temp[0], distorted_height - temp[1]]) for temp in sim_veh])
        sim_veh  = sim_veh * self.car_length / self.car_length_sim
        # sim_veh  = np.int0(sim_veh)
        sim_veh  = sim_veh.tolist()
        # sim_veh  = [tuple([temp[0], temp[1]]) for temp in sim_veh]
        return sim_veh

    def computeSktVehBboxCoord(self, v_color, snap_id):
        skt_veh = self.vehicles[v_color]["snapshots"][snap_id]["min_area_rect"]
        skt_veh = cv2.boxPoints(skt_veh)
        # skt_veh = np.int0(skt_veh)
        skt_veh = skt_veh.tolist()
        # skt_veh = [tuple(temp) for temp in skt_veh]
        return skt_veh

    def midOfVehicle(self, bbox):
        midpoint = ((bbox[0][0] + bbox[1][0] + bbox[2][0] + bbox[3][0]) / len(bbox), 
                    (bbox[0][1] + bbox[1][1] + bbox[2][1] + bbox[3][1]) / len(bbox))  
        return midpoint

    def createRectangle(self, image, box_points, color):
        for i in range(len(box_points) - 1):
            cv2.line(image, box_points[i], box_points[i+1],  color, thickness=2)
        cv2.line(image, box_points[0], box_points[-1],  color, thickness=2)


    """ 
    3. Computing the Bounding Box trajectory Accuracy, IOU and Displacement Error
        a. Accuracy for the BBOX trajectory
        b. IOU (Overlap) Error
        c. Displacement Error
    """
    def computeBboxTrajectory(self, bbox_image, show_image):
        t0 = time.time()
        # bbox_image = image.copy()
        count = 0
        
        for v_color in self.vehicles:
            veh_iou = 0
            displacement = 0
            for snap_id in range(len(self.vehicles[v_color]["snapshots"])):
                skt_veh = self.computeSktVehBboxCoord(v_color, snap_id)
                high_iou_values = list()
                for sim_veh_bbox in self.crash_analysis_log["vehicles"][v_color]["final_bboxes"]:
                    sim_veh = self.computeSimVehBboxCoord(sim_veh_bbox)
                    skt_veh_poly = Polygon(skt_veh)
                    sim_veh_poly = Polygon(sim_veh)
                    iou = skt_veh_poly.intersection(sim_veh_poly).area / skt_veh_poly.union(sim_veh_poly).area
                    # print(iou)
                    if ((iou*100) > 10):
                        high_iou_values.append([iou, skt_veh, sim_veh])
                    else:
                        continue

                high_iou_values = np.array(high_iou_values)
                # print("high IoU values = ", high_iou_values)
                if len(high_iou_values) < 1:
                    continue
                # print("high IoU values = ", high_iou_values[high_iou_values[:,0].argsort()])
                highest_iou = high_iou_values[high_iou_values[:, 0].argsort()][-1]
                # print("Highest IoU = ", highest_iou)
                highest_iou = highest_iou.tolist()
                iou         = highest_iou[0]
                skt_veh     = highest_iou[1]
                sim_veh     = highest_iou[2]
                # print("sketch vehicle midpoint      = ", midOfVehicle(skt_veh))
                # print("simulation vehicle midpoint  = ", midOfVehicle(sim_veh))
                d_error      = self.distanceMetricComputation(np.array(self.midOfVehicle(skt_veh)), np.array(self.midOfVehicle(sim_veh)))
                displacement = displacement + d_error
                # displacement_error.append((True if (d_error    < 5) else False))
                aspect_ratio = self.car_length_sim / self.car_length
                # self.distanceMetricComputation(np.array(init_skt_pos), np.array(init_sim_pos))
                # orientation_distance   = self.distanceMetricComputation(np.array(init_skt_dir), np.array(init_sim_dir))
                # position_similarity    = (True if (position_distance    < 2) else False)

                print("\nVehicle", v_color)
                print("\tVehicle snapshot   =", snap_id)
                print("\tIOU                = {} %".format(round(iou * 100, 2)))
                print("\tDisplacement Error in pixel(s)         =", d_error / aspect_ratio)
                print("\tDisplacement Error in meter(s)         =", d_error)
                print("\tTotal Displacement Error in meter(s)   =", displacement)
                # if ((iou*100) > 50):
                #     veh_iou +=1
                # else:
                #     veh_iou +=0
                
                if snap_id == (len(self.vehicles[v_color]["snapshots"]) - 1):
                    self.crash_analysis_log["vehicles"][v_color]["crash_veh_disp_error"] = d_error
                    self.crash_analysis_log["vehicles"][v_color]["crash_veh_IOU"] = round(iou * 100, 2)
                    self.log["vehicles"][v_color]["crash_veh_disp_error"] = d_error
                    self.log["vehicles"][v_color]["crash_veh_IOU"] = round(iou * 100, 2)
                else:
                    self.crash_analysis_log["vehicles"][v_color]["crash_veh_disp_error"] = "Undefined"
                    self.crash_analysis_log["vehicles"][v_color]["crash_veh_IOU"] = round(iou * 100, 2)
                    self.log["vehicles"][v_color]["crash_veh_disp_error"] = "Undefined"
                    self.log["vehicles"][v_color]["crash_veh_IOU"] = round(iou * 100, 2)
                
                veh_iou += iou * 100
                count   += 20
                cv2.putText(bbox_image, "IoU: {:.2f}%".format(iou * 100), (10, 0 + count), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 255, 0), 2)
                skt_veh = np.int0(skt_veh)
                sim_veh = np.int0(sim_veh)
                sim_veh = [tuple([temp[0], temp[1]]) for temp in sim_veh]
                skt_veh = [tuple(temp) for temp in skt_veh]

                # sketch vehicle BBox
                self.createRectangle(bbox_image, skt_veh, (0, 255, 0))
                # simulator vehicle BBox
                self.createRectangle(bbox_image, sim_veh, (0, 64, 255))

                # for i in range(len(skt_veh)):
                # cv2.circle(bbox_image, skt_veh[i], 3, (0, 255, 0), -1) ## sketch vehicle BBox
                # cv2.circle(bbox_image, sim_veh[i], 3, (0, 128, 255), -1) ## simulator vehicle BBox
                # cv2.imshow("bbox_image", bbox_image)
                if show_image:
                    cv2.imshow("bbox_image", bbox_image)
                    cv2.waitKey(400)
            
            self.crash_analysis_log["vehicles"][v_color]["cum_iou"]             = (veh_iou / (len(self.vehicles[v_color]["snapshots"]) * 100)) * 100
            self.crash_analysis_log["vehicles"][v_color]["cum_iou_error"]       = 100 - self.crash_analysis_log["vehicles"][v_color]["cum_iou"]
            self.crash_analysis_log["vehicles"][v_color]["displacement_error"]  = displacement / len(self.vehicles[v_color]["snapshots"])
            
            #### ---- For Storing  the log in the excel file for data analysis------- ####
            self.log["vehicles"][v_color]["cum_iou"]            = (veh_iou / (len(self.vehicles[v_color]["snapshots"]) * 100)) * 100
            self.log["vehicles"][v_color]["cum_iou_error"]      = 100 - self.crash_analysis_log["vehicles"][v_color]["cum_iou"]
            self.log["vehicles"][v_color]["displacement_error"] = displacement / len(self.vehicles[v_color]["snapshots"])

            print("\nAccuracy for the BBOX trajectory of the {} vehicles = {} % ".format(
                v_color, (veh_iou / (len(self.vehicles[v_color]["snapshots"]) * 100)) * 100))
            print("\nIOU (Overlap) Error =", self.crash_analysis_log["vehicles"][v_color]["cum_iou_error"])
            print("\nDisplacement Error ", displacement / len(self.vehicles[v_color]["snapshots"]))
        print("Overall Average IOU = ", 
                (self.crash_analysis_log["vehicles"]["red"]["cum_iou"] + 
                self.crash_analysis_log["vehicles"]["blue"]["cum_iou"]) / 2)

        cv2.imwrite(self.output_folder + '{}_sim_bbox_traj.jpg'.format(self.process_number), bbox_image)
        self.process_number += 1
        cv2.destroyAllWindows()
        t1 = time.time()
        print("\nBeamNG evaluation -- trajectory of movement in simulation =", t1-t0)
        self.time_efficiency["eval_quality_traj"] = t1-t0



    def computeSimulationAccuracy(self, road_similarity, placement_similarity, orientation_similarity):
        quality_of_environment  = 0.33 * (0.33 * road_similarity + 0.33 * placement_similarity + 0.34 * orientation_similarity)
        quality_of_crash        = 0.34 * (50   * self.crash_analysis_log["vehicles"]["red"]["crash_side_match"] + 
                                          50   * self.crash_analysis_log["vehicles"]["blue"]["crash_side_match"])
        quality_of_trajecory    = 0.33 * (0.5  * self.crash_analysis_log["vehicles"]["red"]["cum_iou"] + 
                                          0.5  * self.crash_analysis_log["vehicles"]["blue"]["cum_iou"])
        simulation_accuracy     = quality_of_environment + quality_of_crash + quality_of_trajecory
        
        print("\n")
        print("Quality_of_environment = {}, quality_of_crash = {}, quality_of_trajecory = {}".format(quality_of_environment, quality_of_crash, quality_of_trajecory))
        print("Crash Simulation Accuracy = ", simulation_accuracy, "%")
        

        #### ---- For Storing  the log in the excel file for data analysis------- ####
        self.log["red_side_match"]      = self.crash_analysis_log["vehicles"]["red"]["crash_side_match"]
        self.log["blue_side_match"]     = self.crash_analysis_log["vehicles"]["blue"]["crash_side_match"]
        self.log["red_cum_iou"]         = self.crash_analysis_log["vehicles"]["red"]["cum_iou"]
        self.log["blue_cum_iou"]        = self.crash_analysis_log["vehicles"]["blue"]["cum_iou"]
        self.log["quality_of_env"]      = quality_of_environment
        self.log["quality_of_crash"]    = quality_of_crash
        self.log["quality_of_traj"]     = quality_of_trajecory
        self.log["simulation_accuracy"] = simulation_accuracy
        
        return simulation_accuracy

    def computeCrisceEfficiency(self):
        print("red cummulative IOU = {}, blue cummulative IOU = {}".format(self.crash_analysis_log["vehicles"]["red"]["cum_iou"], 
                                                                           self.crash_analysis_log["vehicles"]["blue"]["cum_iou"]))
        
        vehicle_ext_time =  self.time_efficiency["preprocess"] + self.time_efficiency["calc_crash_pt"] + self.time_efficiency["seq_movement"] + \
                            self.time_efficiency["tri_ext"] + self.time_efficiency["angle_cal"] + \
                            self.time_efficiency["oriented_nodes"] + self.time_efficiency["skt_veh_impact"]

        roads_ext_time  =   self.time_efficiency["road_ext"] 
        
        traj_ext_time   =   self.time_efficiency["ext_snapshots"] + self.time_efficiency["ext_waypoint"] + \
                            self.time_efficiency["compute_bezier"] + \
                            self.time_efficiency["script_traj"]

        beamng_exe_time =   self.time_efficiency["beamng_loaded"] + self.time_efficiency["beamng_executed"] 
        
        eval_ext_time   =   self.time_efficiency["eval_crash_damage"] + self.time_efficiency["eval_side_impact"] + \
                            self.time_efficiency["eval_quality_road_geom"] + self.time_efficiency["eval_quality_orien_pos"] + \
                            self.time_efficiency["eval_quality_traj"]
        
        total_time      =   {"vehicles_extraction": vehicle_ext_time,
                            "roads_extraction": roads_ext_time,
                            "trajectory_extraction": traj_ext_time,
                            "beamng_execution": beamng_exe_time,
                            "evaluation_time": eval_ext_time}
        
        #### ---- For Storing  the log in the excel file for data analysis------- ####
        self.log["total_time"] = total_time

        return total_time

    def full_extent(self, ax, pad=0.0):
        """Get the full extent of an axes, including axes labels, tick labels, and
        titles."""
        # For text objects, we need to draw the figure first, otherwise the extents
        # are undefined.
        ax.figure.canvas.draw()
        items = ax.get_xticklabels() + ax.get_yticklabels() 
        # items += [ax, ax.title, ax.xaxis.label, ax.yaxis.label]
        items += [ax, ax.title]
        bbox = Bbox.union([item.get_window_extent() for item in items])

        return bbox.expanded(1.1 + pad, 1.1 + pad)


    def plotCrisceEfficiency(self, total_time):
        # fig = plt.figure(figsize=(20,10))
        fig, (ax1, ax2) = plt.subplots(nrows=1, ncols=2, figsize=(16, 8))
        df1 = pd.DataFrame([total_time], columns=['vehicles_extraction', 'roads_extraction', 'trajectory_extraction', 'evaluation_time'])
        df2 = pd.DataFrame([total_time], columns=list(total_time.keys()))

        # Use ax1 to plot Time efficiency of CRISCE
        df1.plot(kind="bar", ax=ax1, title="Time efficiency of CRISCE")
        # ax1.set_title("Time efficiency of CRISCE")
        ax1.set_xticklabels(['values '], rotation="0")
        ax1.set_ylabel("time in seconds")
        extent_1 = self.full_extent(ax1).transformed(fig.dpi_scale_trans.inverted())
        fig.savefig(self.output_folder + '{}_crisce_efficiency.jpg'.format(self.process_number), bbox_inches=extent_1)
        
        # Use ax2 to plot Time efficiency of CRISCE without BeamNG execution
        df2.plot(kind="bar", ax=ax2, title="Time efficiency of CRISCE without BeamNG execution")
        # ax2.set_title("Time efficiency of CRISCE without BeamNG execution")
        ax2.set_xticklabels(['values '], rotation="0")
        ax2.set_ylabel("time in seconds")
        extent_2 = self.full_extent(ax2).transformed(fig.dpi_scale_trans.inverted())
        fig.savefig(self.output_folder + '{}_crisce_beamng_efficiency.jpg'.format(self.process_number), bbox_inches=extent_2)

        # If you don't do tight_layout() you'll have weird overlaps
        plt.tight_layout()
        self.process_number += 1
        # plt.show()
        plt.close()


    def plot_crash(self, vehicle_red, vehicle_blue):
        self.plot_road(plt.gca())
        # self.plot_script(ax[row, 0])
        # self.plot_bbox(plt.gca())
        self.plot_bbox_rect(plt.gca(), vehicle=vehicle_red)
        self.plot_bbox_rect(plt.gca(), vehicle=vehicle_blue)
        # self.plot_overhead(ax[row, 1])

        plt.grid(False)
        plt.axis(False)
        plt.gca().invert_xaxis()
        plt.gca().invert_yaxis()
        # plt.show()  
        plt.close()

    
    def traceVehicleBbox(self):
        Bbox = dict()
        vehicle_red  = self.crash_analysis_log["vehicles"]["red"]["vehicle"]
        vehicle_blue = self.crash_analysis_log["vehicles"]["blue"]["vehicle"]
        Bbox[vehicle_red.vid]  = self.crash_analysis_log["vehicles"]["red"]["final_bboxes"]
        Bbox[vehicle_blue.vid] = self.crash_analysis_log["vehicles"]["blue"]["final_bboxes"]
        # plt.figure(figsize=(30, 10))
        fig_traj, ax = plt.subplots(1, figsize=(30, 20))
        self.plot_road(ax, False)

        for vehicle in Bbox.keys():
            for i, veh_bbox in enumerate(Bbox[vehicle]):
                bbox = veh_bbox
                # boundary_x = veh_bbox[1]
                # boundary_y = veh_bbox[2]
                # bbox = vehicle.get_bbox()
                boundary_x = [
                    bbox['front_bottom_left'][0],
                    bbox['front_bottom_right'][0],
                    bbox['rear_bottom_right'][0],
                    bbox['rear_bottom_left'][0],
                ]
                boundary_y = [
                    bbox['front_bottom_left'][1],
                    bbox['front_bottom_right'][1],
                    bbox['rear_bottom_right'][1],
                    bbox['rear_bottom_left'][1],
                ]

                ax.fill(boundary_x, boundary_y, "r")
                if vehicle == "red_vehicle":
                    ax.fill(boundary_x, boundary_y, 'r-')
                else:
                    ax.fill(boundary_x, boundary_y, 'b-')

                triangle_coord = self.computeTriangle(bbox)
                # print(triangle_coord)
                # ax.Polygon(np.array(triangle_coord), closed=False, color="blue", alpha=0.3, fill=True, edgecolor=None)
                poly = plt.Polygon(np.array(triangle_coord), closed=False,
                                color="white", alpha=1, fill=True, edgecolor=None)
                ax.add_patch(poly)

        # bounding_boxes_red.append(self.getBboxRect(vehicle=vehicle_red)[1:])
        # bounding_boxes_blue.append(self.getBboxRect(vehicle=vehicle_blue)[1:])

        # self.plot_bbox_rect(plt.gca(), vehicle=vehicle_red)
        # self.plot_bbox_rect(plt.gca(), vehicle=vehicle_blue)
        ### plot_overhead(ax[row, 1])

        plt.axis('off')

        plt.grid(False)
        plt.axis(False)
        plt.gca().invert_xaxis()
        plt.gca().invert_yaxis()
        plt.savefig(self.output_folder + '{}_trace_veh_BBOX.jpg'.format(self.process_number), bbox_inches='tight')
        self.process_number += 1
        # plt.show()
        plt.close()
        
        
    def close(self):
        self.bng.close()


    def append_df_to_excel(self, file_path, df, sheet_name='Sheet1', startrow=None,
                        truncate_sheet=False, file_name=None, overwrite=False,
                        **to_excel_kwargs):
        """
        Append a DataFrame [df] to existing Excel file [file_path]
        into [sheet_name] Sheet.
        If [file_path] doesn't exist, then this function will create it.

        @param file_path: File path or existing ExcelWriter
                        (Example: '/path/to/file.xlsx')
        @param df: DataFrame to save to workbook
        @param sheet_name: Name of sheet which will contain DataFrame.
                        (default: 'Sheet1')
        @param startrow: upper left cell row to dump data frame.
                        Per default (startrow=None) calculate the last row
                        in the existing DF and write to the next row...
        @param truncate_sheet: truncate (remove and recreate) [sheet_name]
                            before writing DataFrame to Excel file
        @param to_excel_kwargs: arguments which will be passed to `DataFrame.to_excel()`
                                [can be a dictionary]
        @return: None

        Usage examples:

        >>> append_df_to_excel('d:/temp/test.xlsx', df)

        >>> append_df_to_excel('d:/temp/test.xlsx', df, header=None, index=False)

        >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2',
                            index=False)

        >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2', 
                            index=False, startrow=25)

        (c) [MaxU](https://stackoverflow.com/users/5741205/maxu?tab=profile)
        """

        # Excel file doesn't exist - saving and exiting
        if not os.path.isfile(file_path):
            df.to_excel(
                file_path,
                sheet_name=sheet_name,
                startrow=startrow if startrow is not None else 0,
                header=df.columns.to_list(),
                **to_excel_kwargs)
            # df['file_name'] = df.index
            print("File does not exist creating one !!!!")
            return
            # workbook = xlsxwriter.Workbook(os.path.isfile(file_path))
            # workbook.close()

        print(df.columns.to_list())

        # file_path = file_path.replace("\\", "/")
        if (file_name in pd.read_excel(file_path)["file_name"].to_list() and overwrite == False):
            print("File name already exist !!! ")
            print("For overwriting turn the flag on !!! ")
            return

        elif(file_name in pd.read_excel(file_path)["file_name"].to_list() and overwrite == True):
            print("File name already exist !!! ")
            row_index = df[df.file_name == file_name].index
            print("row_index", row_index)
            # print("File name at index {} is overwritten!!! ".format(row_index))
            df.drop(row_index)

        df.reset_index(drop=True, inplace=True)
            
        print(df.columns.to_list())

        if len(pd.read_excel(file_path)) > 1:
            header = None
        else:
            header = df.columns.to_list()

        # ignore [engine] parameter if it was passed
        if 'engine' in to_excel_kwargs:
            to_excel_kwargs.pop('engine')

        writer = pd.ExcelWriter(file_path, engine='openpyxl', mode='a')

        # try to open an existing workbook
        writer.book = load_workbook(file_path)
        
        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)
        
        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}

        if startrow is None:
            startrow = 0

        # write out the new sheet
        # df['file_name'] = df.index
        df.to_excel(writer, sheet_name, startrow=startrow, header = None, **to_excel_kwargs)

        # save the workbook
        writer.save()





def exec_scenarion(scenario_file):
    RED_CAR_BOUNDARY = np.array([[0, 190, 215],  # red external_0
                            [179, 255, 255]])

    BLUE_CAR_BOUNDARY   = np.array([[85, 50, 60],
                                [160, 255, 255]])
    dir_path = scenario_file
    file = dir_path + "\\sketch.jpeg"
    road = dir_path + "\\road.jpeg"
    external_csv = dir_path + "\\external.csv"
    sketch_type_external = True
    external_impact_points  = None
    
    print("{}\*.csv".format(dir_path))
    if sketch_type_external:
        df = pd.read_csv(external_csv)
        external_impact_points = dict()
        for i in df.index:
            color = str.lower(df.vehicle_color[i])
            impact = str.lower(df.impact_point[i])
            external_impact_points[color] = dict()
            external_impact_points[color] = impact
            
    if sketch_type_external:
        RED_CAR_BOUNDARY = np.array([[0, 190, 215],     # red external crash sketches
                                     [179, 255, 255]])
    else:
        RED_CAR_BOUNDARY = np.array([[0, 200, 180],     # red internal crash sketches
                                     [110, 255, 255]])

    
    show_image = False  # False # True
    # show_image = True  
        
    ########   Main Logic Of the Code Starts Here ################
    
    car = Car()
    roads = Roads()
    kinematics = Kinematics()
    pre_process = Pre_Processing()
    
    sketch = file 
    output_folder = os.path.join(dir_path, "output")  # sketch.split(".")[0])
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    
    # car_length_sim = 5.3
    car_length_sim = 4.670000586694935 # 4.6698
    
    sketch_image_path   = sketch
    road_image_path     = road
    print("\nFile Name = ", sketch)

    #### ------ Read Sketch Image ------ ####### 
    image = pre_process.readImage(sketch_image_path)
    
    car.setColorBoundary(red_boundary=RED_CAR_BOUNDARY, blue_boundary=BLUE_CAR_BOUNDARY)

    vehicles, time_efficiency = car.extractVehicleInformation(image_path=sketch_image_path, time_efficiency=dict(), 
                                                            show_image=show_image, output_folder=output_folder, 
                                                            external=sketch_type_external, external_impact_points=external_impact_points,
                                                            crash_impact_locations=crash_impact_model, car_length_sim=car_length_sim)

    car_length, car_width   = car.getCarDimensions()
    height, width           = car.getImageDimensions()

    roads, lane_nodes = roads.extractRoadInformation(image_path=road_image_path, time_efficiency=time_efficiency, show_image=show_image, 
                                                    output_folder= output_folder, car_length=car_length, car_width=car_width,
                                                    car_length_sim=car_length_sim)

    vehicles, time_efficiency = kinematics.extractKinematicsInformation(image_path=sketch_image_path, vehicles=vehicles, time_efficiency=time_efficiency,
                                                                        output_folder=output_folder, show_image=show_image)
    # print(vehicles)

    simulation_folder = os.path.join(output_folder, "simulation/")
    if not os.path.exists(simulation_folder):
        os.makedirs(simulation_folder)
        
    simulation = Simulation(vehicles=vehicles, roads=roads, 
                            lane_nodes=lane_nodes, kinematics=kinematics, 
                            time_efficiency=time_efficiency, output_folder=simulation_folder,
                            car_length=car_length, car_width=car_width,
                            car_length_sim=car_length_sim, sketch_type_external=sketch_type_external,
                            height=height, width=width)
    
    simulation.bng, simulation.scenario = simulation.setupBeamngSimulation(file)
    simulation.aerialViewCamera()
    
    
    ###### ------ Plotting Roads ---------- ##########
    fig = plt.figure(figsize=(30, 20))
    simulation.plot_road(plt.gca())
    plt.gca().set_aspect("auto")
    # plt.axis('off')
    plt.axis(False)
    plt.gca().invert_yaxis()
    plt.gca().invert_xaxis()
    # plt.show()
    plt.savefig(simulation.output_folder + '{}_sim_plot_road.jpg'.format(simulation.process_number), bbox_inches='tight')
    simulation.process_number += 1
    # fig.savefig(simulation.output_folder + ".jpg", bbox_inches='tight')
    plt.close()
    
    simulation.plotSimulationCrashSketch()
    simulation.initiateDrive()
    simulation.postCrashDamage()
    
    for vehicle_color in vehicles:
        ref_impact_side = vehicles[vehicle_color]["impact_point_details"]["internal_impact_side"]
        print(vehicle_color, ref_impact_side)
        
    simulation.effectAndAccurayOfSimulation()
    
    road_similarity = simulation.computeRoadGeometricSimilarity()
    placement_similarity, orientation_similarity = simulation.computeVehiclesSimilarity()
    simulation.computeBboxTrajectory(image.copy(), show_image=show_image)
    
    simulation_accuracy = simulation.computeSimulationAccuracy(road_similarity, placement_similarity, orientation_similarity)
    
    total_time  = simulation.computeCrisceEfficiency()
    simulation.plotCrisceEfficiency(total_time)
    simulation.traceVehicleBbox()


    simulation.close()
    
    # df = pd.read_excel(dir_path + "\output\output.xlsx") 
    # df.to_json(dir_path + "\output\output.json")
    
    for v_color in vehicles:
        vehicles[v_color]["trajectories"]["computed"]["bezier_curve"]   = [waypoint.tolist() for waypoint in vehicles[v_color]["trajectories"]["computed"]["bezier_curve"]]
        vehicles[v_color]["trajectories"]["computed"]["b_spline"]       = [waypoint.tolist() for waypoint in vehicles[v_color]["trajectories"]["computed"]["b_spline"]]
        vehicles[v_color]["trajectories"]["computed"]["cubic_spline"]   = [waypoint.tolist() for waypoint in vehicles[v_color]["trajectories"]["computed"]["cubic_spline"]]
        vehicles[v_color]["trajectories"]["original_trajectory"]    = [waypoint.tolist() for waypoint in vehicles[v_color]["trajectories"]["original_trajectory"]]
        vehicles[v_color]["trajectories"]["distorted_trajectory"]   = [waypoint.tolist() for waypoint in vehicles[v_color]["trajectories"]["distorted_trajectory"]]
        vehicles[v_color]["trajectories"]["simulation_trajectory"]  = [waypoint.tolist() for waypoint in vehicles[v_color]["trajectories"]["simulation_trajectory"]]


    log = {"vehicles": vehicles,
            "road_nodes": lane_nodes,
            "vehicle_crash_specifics": simulation.log["vehicles"],
            "simulation_trajectory":{"red": simulation.crash_analysis_log["vehicles"]["red"]["simulation_trajectory"],
                                     "blue": simulation.crash_analysis_log["vehicles"]["blue"]["simulation_trajectory"]},
            "simulation_rotations": {"red": simulation.crash_analysis_log["vehicles"]["red"]["rotations"],
                                     "blue": simulation.crash_analysis_log["vehicles"]["blue"]["rotations"]},
            "sim_veh_speed": {"red": simulation.crash_analysis_log["vehicles"]["red"]["wheel_speed"],
                              "blue": simulation.crash_analysis_log["vehicles"]["blue"]["wheel_speed"]},
            "sim_time": {"red": simulation.crash_analysis_log["vehicles"]["red"]["sim_time"],
                         "blue": simulation.crash_analysis_log["vehicles"]["blue"]["sim_time"]},
            "simulated_impact": simulation.log["simulated_impact"], 
            "road_similarity": simulation.log["road_similarity"],
            "placement_similarity" : simulation.log["placement_similarity"], 
            "orientation_similarity": simulation.log["orientation_similarity"],
            "quality_of_env": simulation.log["quality_of_env"], "red_side_match" : simulation.log["red_side_match"], 
            "blue_side_match": simulation.log["blue_side_match"], "quality_of_crash": simulation.log["quality_of_crash"], 
            "red_cum_iou" : simulation.log["red_cum_iou"], "blue_cum_iou": simulation.log["blue_cum_iou"], 
            "quality_of_traj": simulation.log["quality_of_traj"], 
            "crash_disp_error": {"red": simulation.log["vehicles"]["red"]["crash_veh_disp_error"],
                                 "blue": simulation.log["vehicles"]["blue"]["crash_veh_disp_error"]},
            "crash_IOU": {"red": simulation.log["vehicles"]["red"]["crash_veh_IOU"],
                          "blue": simulation.log["vehicles"]["blue"]["crash_veh_IOU"]},
            "simulation_accuracy": simulation.log["simulation_accuracy"], 
            "total_time": simulation.log["total_time"]}
    
    
    with open(dir_path + '\output.json', 'w') as fp:
        json.dump(log, fp,  indent=1)
        
    roads = []
    for i, road_nodes in enumerate(lane_nodes):
        road_dict = {
            "name":"road"+str(i+1),
            "road_type":"roadway",
            "road_shape":"I",
            "road_node_list":road_nodes
        }
        roads.append(road_dict)
    
    vehicle_list = []
    def angle_to_quat(angle):
        import numpy as np
        angle = np.radians(angle)

        cy = np.cos(angle[2] * 0.5)
        sy = np.sin(angle[2] * 0.5)
        cp = np.cos(angle[1] * 0.5)
        sp = np.sin(angle[1] * 0.5)
        cr = np.cos(angle[0] * 0.5)
        sr = np.sin(angle[0] * 0.5)

        w = cr * cp * cy + sr * sp * sy
        x = sr * cp * cy - cr * sp * sy
        y = cr * sp * cy + sr * cp * sy
        z = cr * cp * sy - sr * sp * cy

        return (x, y, z, w)
    for i, v_color in enumerate(["red", "blue"]):
        veh_dict = {
            "name":"v"+str(i+1),
            "color":"1 0 0" if v_color=="red" else "0 0 1",
            "rot_quat":angle_to_quat(simulation.crash_analysis_log["vehicles"][v_color]["rot"]),
            "distance_to_trigger":-1.0,
            "damage_components":"ANY"
        }
        speed = max(log["sim_veh_speed"][v_color])
        
        debug_trajectory = vehicles[v_color]["trajectories"]["debug_trajectory"]
        driving_actions = []
        idx = 0
        while (idx+1 < len(debug_trajectory)):
            driving_actions.append({
                    "name":"follow",
                    "trajectory":[[debug_trajectory[idx], debug_trajectory[idx+1]]],
                    "speed":speed
            })
            idx+=1
        if len(debug_trajectory) == 1:
            driving_actions.append({
                    "name":"stop",
                    "trajectory":[[debug_trajectory[0]]],
                    "speed":0
            })
        veh_dict["driving_actions"] = driving_actions
        vehicle_list.append(veh_dict)
        
    expected_crash_components = []
    for i, v_color in enumerate(["red", "blue"]):
        ref_parts = simulation.log["vehicles"][v_color]["ref_side_values"]
        part_names = set()
        for part in ref_parts:
            part_name = ""
            for key in crash_impact_model:
                if part in crash_impact_model[key]:
                    part_name = key.split("_")[0] + ' ' + key.split("_")[1]
                    if part_name == "front mid":
                        part_name = "front"
                    if part_name == "rear mid":
                        part_name = "rear"
                    if part_name == "left mid":
                        part_name = "middle left"
                    if part_name == "right mid":
                        part_name = "middle right"
            part_names.add(part_name)
        parts = []
        for name in part_names:
            parts.append({
                "name": name,
                "damage": 1
            })
    
        crash_dict = {
            "name":"v"+str(i+1),
            "parts": parts
        }
        expected_crash_components.append(crash_dict)
        
    ac3r_plus = {
        "name": scenario_file.split("\\")[3],
        "roads": roads,
        "vehicles": vehicle_list,
        "expected_crash_components": expected_crash_components
    }
    
    path_ac3rp = "crisce" + '\\' + scenario_file.split("\\")[3] + "_data.json"
    with open(path_ac3rp, 'w') as fp:
        json.dump(ac3r_plus, fp)

    print("END")


@click.group()
def cli():
    pass

@cli.command()
@click.argument('scenario_file')
def run_from_scenario(scenario_file):
    exec_scenarion(scenario_file)
    


if __name__ == '__main__':
    import sys
    my_dict = {
        0: '.\\Datasets\\NMVCCS\\2005002585724',
        # 1: '.\\Datasets\\NMVCCS\\2005008586061a',
        # 2: '.\\Datasets\\NMVCCS\\2005008586061b',
        # 3: '.\\Datasets\\NMVCCS\\2005012695622',
        # 4: '.\\Datasets\\NMVCCS\\2005045587341',
        # 5: '.\\Datasets\\NMVCCS\\2005048103904',
        # 6: '.\\Datasets\\NMVCCS\\2006048103049',
        # 7: '.\\Datasets\\CIREN\\100343',
        # 8: '.\\Datasets\\CIREN\\156722',
    }
    for k in my_dict:
        exec_scenarion(my_dict[k])

    